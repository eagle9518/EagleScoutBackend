import openpyxl
import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from setuptools import glob
from string import *

HEADERS = ["Team", "Match_Num", "Auto_Cross", "Auto_Upper",
           "Auto_Bottom", "Tele_Upper", "Tele_Bottom", "Level",
           "Driver_Performance", "Defense_Perf", "Name", "Comments"]

darkRedFill = PatternFill(start_color='FF3333', end_color='FF3333', fill_type='solid')
redFill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
orangeFill = PatternFill(start_color='FFB266', end_color='FFB266', fill_type='solid')
yellowFill = PatternFill(start_color='FFFF66', end_color='FFFF66', fill_type='solid')
limeFill = PatternFill(start_color='B2FF66', end_color='B2FF66', fill_type='solid')
greenFill = PatternFill(start_color='66FF66', end_color='66FF66', fill_type='solid')

light_green_fill = PatternFill("solid", start_color="16e064")
dark_green_fill = PatternFill("solid", start_color="078f3b")

def CSV_Reader():
    return pd.concat([(pd.read_csv(file, names=HEADERS)) for file in glob.glob('New_CSVs/*.csv')], ignore_index=True)

def teams_writer():
    combinedData = CSV_Reader()
    teams = combinedData.groupby("Team")
    teamData = dict(tuple(teams))
    with pd.ExcelWriter('Excel_Sheets/Teams.xlsx') as writer:
        for team in sorted(teamData.keys()):
	    try:
            	teamData[team].to_excel(writer, sheet_name="Team" + str(team), index=False)
	    except Exception as e:
		print(team, e)
def main():
    teams_writer()
    # rankings_writer()
    nishan_update()

def nishan_update():
    """
    separate column for auto score
    separate column for tele score
    Stats for Average, Total, stdev, stdev/avg
    """
    workbook = openpyxl.load_workbook("Excel_Sheets/Teams.xlsx")
    workbook.create_sheet("Rankings")
    ranking_sheet = workbook["Rankings"]

    ranking_sheet["A1"] = "Team"
    ranking_sheet["B1"] = "Auto"
    ranking_sheet["E1"] = "Teleop"
    ranking_sheet["H1"] = "Climb"
    ranking_sheet["K1"] = "Average Points"
    ranking_sheet["N1"] = "Driver Perf"
    ranking_sheet["Q1"] = "Defense Perf"

    for i in range(1, 19, 3):
        ranking_sheet["%s2"%(ascii_uppercase[i])] = "Mean"
        ranking_sheet["%s2"%(ascii_uppercase[i+1])] = "STDev"
        ranking_sheet["%s2"%(ascii_uppercase[i+2])] = "STDev % of mean"

    for team, sheet in enumerate(workbook.worksheets[:-1]):
        if sheet.cell(row=6, column=1) != "Auto_Score":

            num_rows = get_num_rows(sheet)
            sheet.insert_cols(6)
            sheet["F1"] = "Total Auto"
            sheet["F1"].fill = light_green_fill

            sheet.insert_cols(9)
            sheet["I1"] = "Total Teleop"
            sheet["I1"].fill = light_green_fill

            sheet.insert_cols(11)
            sheet["K1"] = "Climb_Score"
            sheet["K1"].fill = light_green_fill

            sheet.insert_cols(12)
            sheet["L1"] = "Total Points"
            sheet["L1"].fill = dark_green_fill

            for i in range(num_rows-2):
                sheet["F%s"%(i+2)] = "=C%s*2+D%s*4+E%s*2"%(i+2, i+2, i+2)
                sheet["F%s"%(i+2)].fill = light_green_fill
                
                sheet["I%s"%(i+2)] = "=G%s*2+H%s"%(i+2, i+2)
                sheet["I%s"%(i+2)].fill = light_green_fill

                sheet["K%s"%(i+2)] = "=IF(J%s = 0, 0, IF(J%s = 1, 4, IF(J%s = 2, 6, IF(J%s = 3, 10, IF(J%s = 4, 15, 0)))))"%(i+2, i+2, i+2, i+2, i+2)
                sheet["K%s"%(i+2)].fill = light_green_fill

                sheet["L%s"%(i+2)] = "=F%s+I%s+K%s"%(i+2, i+2, i+2)
                sheet["L%s"%(i+2)].fill = dark_green_fill

            sheet["B20"] = "Average:"
            sheet["B21"] = "STDev:"
            sheet["B22"] = "stDEV % of Average:"

            for letter in ascii_uppercase[2:14]:
                sheet["%s20"%(letter)] = "=AVERAGE(%s2:%s14)"%(letter, letter)
                sheet["%s21"%(letter)] = "=STDEV(%s2:%s14)"%(letter, letter)
                sheet["%s22"%(letter)] = "=%s21/%s20"%(letter, letter)

            # Writing to Rankings Sheet
            team_number = sheet.cell(row=2, column=1).value
            ranking_sheet["A%s"%(team+3)] = team_number

            rankingsSheetCorrespond = {1:'F', 4:'I', 7:'K', 10:'L', 13:'M', 16:'N'}
            for i in rankingsSheetCorrespond.keys():
                ranking_sheet["%s%s"%(ascii_uppercase[i], team+3)] = "=Team%s!%s20"%(team_number, rankingsSheetCorrespond[i])
                ranking_sheet["%s%s"%(ascii_uppercase[i+1], team+3)] = "=Team%s!%s21"%(team_number, rankingsSheetCorrespond[i])
                ranking_sheet["%s%s"%(ascii_uppercase[i+2], team+3)] = "=Team%s!%s22"%(team_number, rankingsSheetCorrespond[i])
                
    workbook.save("Excel_Sheets/Teams.xlsx")


def get_num_rows(ws):
    curr_row = 1
    while True:
        if ws.cell(row=curr_row, column=1).value is not None:
            curr_row += 1
        else:
            break
    return curr_row 

if __name__ == '__main__':
    main()

